import csv
import dataclasses
import logging
import pathlib
import re

import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.worksheet

from .models import LocalisationData, LocFile, LocLine, Text, TranslationData, TranslationEntry, TranslationStatus

_LOC_LANG_RE = re.compile(r"^l_([a-z]+):$")
_LOC_SEPARATOR_RE = re.compile(r":[0-9]")


@dataclasses.dataclass
class ReloadStats:
    new: int
    changed: int
    deleted: int

    outdated_translations: int

    @property
    def all(self) -> int:
        return self.new + self.changed + self.deleted


def _load_loc_from_file(
    filepath: pathlib.Path,
    language: str,
) -> LocFile | None:
    logging.info(f"Parsing locfile {str(filepath)!r}")
    lines = []
    with open(filepath, "r", encoding="utf-8-sig") as fh:
        # Get the language
        file_language_raw = fh.readline().strip()
        file_language_match = _LOC_LANG_RE.match(file_language_raw)
        if not file_language_match:
            logging.warning(f"Could not find language from file {filepath.name!r}")
            return None
        file_language = file_language_match.group(1)
        # Ignore unwanted languages
        if file_language != language:
            logging.info(f"Skipping {filepath.name!r}, wrong language ({file_language!r} instead of {language!r})")
            return None
        # Parse lines
        for line_nr, line in enumerate(fh, start=2):
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            try:
                identifier, text = _LOC_SEPARATOR_RE.split(line, maxsplit=1)
            except ValueError:
                logging.warning(f"Invalid entry in localisation file {filepath.name!r}, line {line_nr}")
                continue
            else:
                text = text.strip().removeprefix('"').removesuffix('"').replace('\\"', '"')
                lines.append(
                    LocLine(
                        identifier=identifier.strip(),
                        text=text.strip(),
                    )
                )
    return LocFile(sourcefile=filepath, language=language, lines=lines)


def _merge_localisations(
    locfiles: list[LocFile],
    language: str,
) -> LocalisationData:
    logging.info("Merging localisations")
    locdata = LocalisationData(language=language)
    for locfile in locfiles:
        if locfile.language != locdata.language:
            continue
        for locline in locfile.lines:
            if locline.identifier in locdata.entries:
                logging.warning(f"Skipping duplicate definition of {locline.identifier!r}")
                continue
            locdata.entries[locline.identifier] = locline.text
    return locdata


def parse_localisation_from_locfiles(
    filepaths: list[pathlib.Path],
    language: str,
) -> LocalisationData:
    locfiles = []
    for filepath in filepaths:
        if not filepath.exists():
            continue
        locfile = _load_loc_from_file(filepath=filepath, language=language)
        if locfile is not None:
            locfiles.append(locfile)
    return _merge_localisations(
        locfiles=locfiles,
        language=language,
    )


def parse_translations_from_excel(filepath: pathlib.Path) -> TranslationData:
    logging.info(f"Parsing Excel {str(filepath)!r}")
    wb = openpyxl.load_workbook(filepath)
    ws: openpyxl.worksheet.worksheet.Worksheet = wb["translations"]
    # Read header
    translation_language = ws.cell(row=1, column=3).value
    reference_language = ws.cell(row=1, column=4).value
    locdata = TranslationData(
        reference_language=reference_language,
        translation_language=translation_language,
    )
    # Read rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
        values = [str(v or "") for v in row]  # ENsure all are strings
        identifier, raw_status, translation, reference = values
        status = TranslationStatus(
            raw_status or (TranslationStatus.DONE.value if translation else TranslationStatus.MISSING.value)
        )
        locdata.entries[identifier] = TranslationEntry(
            reference=reference,
            translation=translation,
            status=status,
        )
    return locdata


def write_localisation_to_locfile(
    outfile: pathlib.Path,
    locdata: TranslationData,
) -> int:
    logging.info(f"Writing localisation for language {locdata.language!r} to {str(outfile)!r}")
    written = 0
    with open(outfile, "w", encoding="utf-8") as fh:
        fh.write(f"l_{locdata.language}:\n")
        for identifier, text in locdata.entries.items():
            if text:
                text = text.replace('"', '\\"')
                fh.write(f' {identifier}:0 "{text}"\n')
                written += 1
    return written


def write_translations_to_excel(
    outpath: pathlib.Path,
    translation_data: TranslationData,
):
    if outpath.exists():
        wb = openpyxl.load_workbook(outpath, data_only=True)
    else:
        wb = openpyxl.Workbook()
    for sheetname in wb.sheetnames:
        del wb[sheetname]
    ws: openpyxl.worksheet.worksheet.Worksheet = wb.create_sheet(title="translations")
    # Create header row
    for colnr, colname in enumerate(
        [
            "identifier",
            "translation_status",
            translation_data.translation_language,
            translation_data.reference_language,
        ],
        start=1,
    ):
        ws.cell(row=1, column=colnr, value=colname)
    # Write translations
    for rownr, (locid, entry) in enumerate(translation_data.entries.items(), start=2):
        status = entry.status.value if entry.status is TranslationStatus.OUTDATED else ""
        ws.cell(row=rownr, column=1, value=locid)
        ws.cell(row=rownr, column=2, value=status)
        ws.cell(row=rownr, column=3, value=entry.translation)
        ws.cell(row=rownr, column=4, value=entry.reference)
    wb.save(outpath)


def merge_latest_references_into_translations(
    known_translations: TranslationData,
    latest_locdata: LocalisationData,
) -> tuple[TranslationData, ReloadStats]:
    updated_translations = TranslationData(
        reference_language=known_translations.reference_language,
        translation_language=known_translations.translation_language,
    )
    all_locids = known_translations.entries.keys() | latest_locdata.entries.keys()
    stats = ReloadStats(0, 0, 0, 0)
    for locid in all_locids:
        latest_reference = latest_locdata.entries.get(locid, "")
        current_entry = known_translations.entries.get(locid, TranslationEntry("", "", TranslationStatus.MISSING))
        new_status = _determine_status(
            current_status=current_entry.status,
            prev_reference=current_entry.reference,
            new_reference=latest_reference,
        )
        updated_translations.entries[locid] = TranslationEntry(
            reference=latest_reference,
            translation=current_entry.translation,
            status=new_status,
        )
        if locid not in latest_locdata.entries:
            stats.deleted += 1
        elif locid not in known_translations.entries:
            stats.new += 1
        elif latest_reference != current_entry.reference:
            stats.changed += 1
        if new_status != current_entry.status:
            stats.outdated_translations += 1
    return updated_translations, stats


def _determine_status(
    current_status: TranslationStatus,
    prev_reference: Text,
    new_reference: Text,
) -> TranslationStatus:
    if current_status is TranslationStatus.DONE:
        if prev_reference == new_reference:
            return TranslationStatus.DONE
        else:
            return TranslationStatus.OUTDATED
    else:
        return current_status


def get_localisation_from_translations(translation_data: TranslationData) -> LocalisationData:
    locdata = LocalisationData(language=translation_data.translation_language)
    for locid, entry in translation_data.entries.items():
        locdata.entries[locid] = entry.translation
    return locdata
